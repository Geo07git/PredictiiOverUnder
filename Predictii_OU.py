import hashlib
import json
import os
import pickle
import warnings
from datetime import datetime
from io import StringIO

import numpy as np
import pandas as pd
import plotly.graph_objects as go
import requests
import streamlit as st
from scipy.stats import poisson
from sklearn.calibration import CalibratedClassifierCV, calibration_curve
from sklearn.inspection import permutation_importance
from sklearn.metrics import brier_score_loss, roc_auc_score, log_loss
from sklearn.model_selection import TimeSeriesSplit
from xgboost import XGBClassifier

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

#streamlit run over_predictor_v4_2.py

warnings.filterwarnings("ignore")

st.set_page_config(
    page_title="Predictii O/U",
    page_icon="⚽",
    layout="wide",
    initial_sidebar_state="expanded",
)

CONDITIONS = {
    "Over/Under 1.5": {"threshold": 1.5, "label": "O/U 1.5", "emoji": "1️⃣"},
    "Over/Under 2.5": {"threshold": 2.5, "label": "O/U 2.5", "emoji": "2️⃣"},
    "Over/Under 3.5": {"threshold": 3.5, "label": "O/U 3.5", "emoji": "3️⃣"},
}

MODEL_PATHS = {
    "Over/Under 1.5": "model_over15_v4_2.pkl",
    "Over/Under 2.5": "model_over25_v4_2.pkl",
    "Over/Under 3.5": "model_over35_v4_2.pkl",
}

DATA_PATH = "date_over_v4_2.pkl"
HISTORY_PATH = "Predictii_OU.xlsx"
HISTORY_SHEET = "Istoric"
HISTORY_TABLE = "tblPredictiiOU"
LIGI = {
    "ENG-Premier League": "https://www.football-data.co.uk/mmz4281/{s}/E0.csv",
    "ENG-Championship": "https://www.football-data.co.uk/mmz4281/{s}/E1.csv",
    "ENG-League One": "https://www.football-data.co.uk/mmz4281/{s}/E2.csv",
    "ENG-League Two": "https://www.football-data.co.uk/mmz4281/{s}/E3.csv",
    "ENG-League Conf": "https://www.football-data.co.uk/mmz4281/{s}/EC.csv",
    "ESP-La Liga": "https://www.football-data.co.uk/mmz4281/{s}/SP1.csv",
    "ESP-La Liga 2": "https://www.football-data.co.uk/mmz4281/{s}/SP2.csv",
    "GER-Bundesliga": "https://www.football-data.co.uk/mmz4281/{s}/D1.csv",
    "GER-Bundesliga 2": "https://www.football-data.co.uk/mmz4281/{s}/D2.csv",
    "ITA-Serie A": "https://www.football-data.co.uk/mmz4281/{s}/I1.csv",
    "ITA-Serie B": "https://www.football-data.co.uk/mmz4281/{s}/I2.csv",
    "FRA-Ligue 1": "https://www.football-data.co.uk/mmz4281/{s}/F1.csv",
    "FRA-Ligue 2": "https://www.football-data.co.uk/mmz4281/{s}/F2.csv",
    "NED-Eredivisie": "https://www.football-data.co.uk/mmz4281/{s}/N1.csv",
    "POR-Primeira Liga": "https://www.football-data.co.uk/mmz4281/{s}/P1.csv",
    "BEL-Jupiler Pro": "https://www.football-data.co.uk/mmz4281/{s}/B1.csv",
    "TUR-Super Lig": "https://www.football-data.co.uk/mmz4281/{s}/T1.csv",
    "GRE-Super League": "https://www.football-data.co.uk/mmz4281/{s}/G1.csv",
    "SCO-Scottish Prem": "https://www.football-data.co.uk/mmz4281/{s}/SC0.csv",
    "SCO-Scottish Champ": "https://www.football-data.co.uk/mmz4281/{s}/SC1.csv",
    "SCO-Scottish Div1": "https://www.football-data.co.uk/mmz4281/{s}/SC2.csv",
    "SCO-Scottish Div2": "https://www.football-data.co.uk/mmz4281/{s}/SC3.csv",
}

SEZOANE = ["1617", "1718", "1819", "1920", "2021", "2122", "2223", "2324", "2425", "2526"]

FEATURES_COMMON = [
    "lambda_home", "lambda_away", "lambda_total",
    "home_xg_proxy", "away_xg_proxy", "xg_diff",
    "home_xg_conceded_proxy", "away_xg_conceded_proxy",
    "home_sot", "away_sot",
    "home_sot_conceded", "away_sot_conceded",
    "home_conversion", "away_conversion", "conversion_diff",
    "sot_pct_home", "sot_pct_away",
    "home_goals_scored", "away_goals_scored",
    "home_goals_conceded", "away_goals_conceded",
    "home_advantage",
    "home_form5", "away_form5",
    "home_form10", "away_form10",
    "home_corners", "away_corners", "corners_total",
    "league_tempo",
    "home_fouls", "away_fouls",
    "home_yellows", "away_yellows",
    "elo_home", "elo_away", "elo_diff", "elo_diff_norm",
    "attack_h_vs_def_a", "attack_a_vs_def_h",
    "btts_proxy", "strength_diff",
    "season_progress",
    "home_days_rest", "away_days_rest", "rest_diff",
    "league_over_mean",
    "home_over_rate", "away_over_rate",
    "home_over_rate_adj", "away_over_rate_adj",
    "poisson_over",
]

FEATURES_EXTRA = {
    "Over/Under 1.5": [
        "home_goals_scored", "away_goals_scored", "home_form5", "away_form5",
        "elo_diff_norm", "home_sot", "away_sot", "home_conversion", "away_conversion",
    ],
    "Over/Under 2.5": [
        "poisson_over", "home_xg_proxy", "away_xg_proxy", "home_over_rate",
        "away_over_rate", "btts_proxy", "lambda_total", "corners_total",
    ],
    "Over/Under 3.5": [
        "lambda_total", "lambda_home", "lambda_away", "attack_h_vs_def_a",
        "attack_a_vs_def_h", "corners_total", "league_tempo", "home_xg_proxy",
        "away_xg_proxy", "home_sot", "away_sot", "btts_proxy", "strength_diff",
    ],
}

DEFAULT_PARAMS = {
    "Over/Under 1.5": {
        "n_estimators": 500, "max_depth": 4, "learning_rate": 0.04,
        "subsample": 0.8, "colsample_bytree": 0.7,
        "reg_alpha": 1.0, "reg_lambda": 3.0,
        "min_child_weight": 10, "gamma": 0.2,
    },
    "Over/Under 2.5": {
        "n_estimators": 600, "max_depth": 5, "learning_rate": 0.05,
        "subsample": 0.8, "colsample_bytree": 0.7,
        "reg_alpha": 0.8, "reg_lambda": 2.5,
        "min_child_weight": 8, "gamma": 0.15,
    },
    "Over/Under 3.5": {
        "n_estimators": 700, "max_depth": 5, "learning_rate": 0.04,
        "subsample": 0.75, "colsample_bytree": 0.65,
        "reg_alpha": 1.2, "reg_lambda": 3.5,
        "min_child_weight": 12, "gamma": 0.25,
    },
}

st.markdown(
    """
<style>
    .badge {display:inline-block;padding:4px 10px;border-radius:999px;font-size:12px;font-weight:700;margin:2px 4px 6px 0;}
    .badge-green{background:#d1e7dd;color:#0f5132;}
    .badge-orange{background:#fff3cd;color:#664d03;}
    .badge-red{background:#f8d7da;color:#842029;}
    .badge-blue{background:#cfe2ff;color:#084298;}
    .condition-badge{display:inline-block;padding:6px 16px;border-radius:999px;font-weight:700;margin-bottom:8px;}
    .cond-15{background:#cfe2ff;color:#084298;}
    .cond-25{background:#d1e7dd;color:#0f5132;}
    .cond-35{background:#f8d7da;color:#842029;}
    .info-box{background:#eef6ff;border-left:4px solid #4f8df5;padding:.75rem 1rem;border-radius:0;margin-bottom:1rem;}
    .warn-box{background:#fff4e5;border-left:4px solid #ffb74d;padding:.75rem 1rem;border-radius:0;margin-bottom:1rem;}
</style>
""",
    unsafe_allow_html=True,
)


def df_hash(df: pd.DataFrame) -> str:
    try:
        payload = pd.util.hash_pandas_object(df, index=False).values.tobytes()
        return hashlib.md5(payload).hexdigest()
    except Exception:
        return hashlib.md5(str((len(df), tuple(df.columns))).encode()).hexdigest()


def sfloat(v, default=np.nan):
    try:
        fv = float(v)
        return default if np.isnan(fv) else fv
    except Exception:
        return default


def rolling_ewm(df_sorted: pd.DataFrame, echipa_col: str, val_col: str, decay: float = 0.05) -> pd.Series:
    span = max(1, int(2.0 / decay - 1))
    result = pd.Series(np.nan, index=df_sorted.index, dtype=float)
    if val_col not in df_sorted.columns:
        return result
    for _, grp in df_sorted.groupby([echipa_col, "Liga"], sort=False, observed=True):
        grp = grp.sort_values("Date")
        shifted = grp[val_col].shift(1)
        ewm_vals = shifted.ewm(span=span, min_periods=3, adjust=True).mean()
        result.loc[grp.index] = ewm_vals.values
    return result


def poisson_over(lh: float, la: float, threshold: float) -> float:
    if np.isnan(lh) or np.isnan(la) or lh <= 0 or la <= 0:
        return np.nan
    int_thresh = int(threshold)
    p_under = 0.0
    for gh in range(int_thresh + 1):
        for ga in range(int_thresh + 1 - gh):
            p_under += poisson.pmf(gh, lh) * poisson.pmf(ga, la)
    return float(1.0 - p_under)


def poisson_over_series(lh_arr, la_arr, threshold: float) -> np.ndarray:
    int_thresh = int(threshold)
    lh_arr = np.asarray(lh_arr, dtype=float)
    la_arr = np.asarray(la_arr, dtype=float)
    result = np.full(len(lh_arr), np.nan)
    valid = (~np.isnan(lh_arr)) & (~np.isnan(la_arr)) & (lh_arr > 0) & (la_arr > 0)
    if not valid.any():
        return result
    lh_v = lh_arr[valid]
    la_v = la_arr[valid]
    p_under = np.zeros(valid.sum())
    for gh in range(int_thresh + 1):
        for ga in range(int_thresh + 1 - gh):
            p_under += poisson.pmf(gh, lh_v) * poisson.pmf(ga, la_v)
    result[valid] = 1.0 - p_under
    return result


def elo_update(elo_h, elo_a, gols_h, gols_a, k=28.0):
    expected_h = 1.0 / (1.0 + 10 ** ((elo_a - elo_h) / 400.0))
    result_h = 1.0 if gols_h > gols_a else (0.5 if gols_h == gols_a else 0.0)
    delta = k * (result_h - expected_h)
    return elo_h + delta, elo_a - delta


def calculeaza_elo_serii(df: pd.DataFrame) -> pd.DataFrame:
    df_sorted = df.sort_values("Date").reset_index(drop=True)
    n = len(df_sorted)
    elo_ratings = {}
    elo_h_arr = np.full(n, 1500.0)
    elo_a_arr = np.full(n, 1500.0)
    home_teams = df_sorted["HomeTeam"].values
    away_teams = df_sorted["AwayTeam"].values
    fthg = df_sorted["FTHG"].values
    ftag = df_sorted["FTAG"].values
    for i in range(n):
        h, a = home_teams[i], away_teams[i]
        eh = elo_ratings.get(h, 1500.0)
        ea = elo_ratings.get(a, 1500.0)
        elo_h_arr[i] = eh
        elo_a_arr[i] = ea
        try:
            new_eh, new_ea = elo_update(eh, ea, int(fthg[i]), int(ftag[i]))
            elo_ratings[h] = new_eh
            elo_ratings[a] = new_ea
        except Exception:
            pass
    df_sorted["elo_home"] = elo_h_arr
    df_sorted["elo_away"] = elo_a_arr
    df_sorted["elo_diff"] = elo_h_arr - elo_a_arr
    return df_sorted.set_index(df.sort_values("Date").index)


def bootstrap_ci(prob: float, n: int = 500, ci: float = 0.90):
    alpha = (1 - ci) / 2
    samples = np.random.binomial(1, prob, size=(n, 100)).mean(axis=1)
    return float(np.quantile(samples, alpha)), float(np.quantile(samples, 1 - alpha))


def get_features_for_condition(condition: str, df_cols: list) -> list:
    all_feats = list(dict.fromkeys(FEATURES_COMMON + FEATURES_EXTRA.get(condition, [])))
    return [f for f in all_feats if f in df_cols]


def load_all_models():
    models = {}
    for cond, path in MODEL_PATHS.items():
        if os.path.exists(path):
            try:
                with open(path, "rb") as f:
                    models[cond] = pickle.load(f)
            except Exception:
                models[cond] = None
        else:
            models[cond] = None
    return models


def descarca_fdco(progress_bar, status_text) -> pd.DataFrame:
    toate = []
    total = len(LIGI) * len(SEZOANE)
    i = 0
    for liga, url_tmpl in LIGI.items():
        for sezon in SEZOANE:
            url = url_tmpl.format(s=sezon)
            try:
                r = requests.get(url, timeout=15)
                if r.status_code == 200 and len(r.content) > 500:
                    df = pd.read_csv(StringIO(r.text), low_memory=False)
                    df["Liga"] = liga
                    df["Sezon"] = sezon
                    toate.append(df)
                    status_text.text(f"✅ {liga} {sezon} — {len(df)} meciuri")
            except Exception as e:
                status_text.text(f"⚠️ {liga} {sezon}: {e}")
            i += 1
            progress_bar.progress(i / total)
    return pd.concat(toate, ignore_index=True) if toate else pd.DataFrame()


@st.cache_data(show_spinner=False, ttl=3600)
def proceseaza_date_cached(data_hash: str, df_raw: pd.DataFrame) -> pd.DataFrame:
    return _proceseaza_date_core(df_raw)


def _days_rest(df: pd.DataFrame, team_col: str) -> pd.Series:
    rezultate = {}
    for _, grp in df.groupby(team_col, observed=True):
        grp = grp.sort_values("Date")
        diffs = grp["Date"].diff().dt.days.fillna(21)
        rezultate.update(dict(zip(grp.index, diffs)))
    return pd.Series(rezultate)


def _proceseaza_date_core(df_raw: pd.DataFrame) -> pd.DataFrame:
    cols_necesare = ["Date", "HomeTeam", "AwayTeam", "FTHG", "FTAG", "Liga", "Sezon"]
    cols_opt = ["HST", "AST", "HS", "AS", "HC", "AC", "HF", "AF", "HY", "AY"]
    keep = [c for c in cols_necesare + cols_opt if c in df_raw.columns]
    df = df_raw[keep].copy()
    df["Date"] = pd.to_datetime(df["Date"], dayfirst=True, errors="coerce")
    for col in [c for c in ["FTHG", "FTAG"] + cols_opt if c in df.columns]:
        df[col] = pd.to_numeric(df[col], errors="coerce")
    df = df.dropna(subset=["Date", "FTHG", "FTAG", "HomeTeam", "AwayTeam", "Liga", "Sezon"]).copy()
    df["total_goals"] = df["FTHG"] + df["FTAG"]
    for col in ["Liga", "Sezon", "HomeTeam", "AwayTeam"]:
        df[col] = df[col].astype("category")
    df = df.sort_values(["Liga", "Date"]).reset_index(drop=True)

    elo_parts = [calculeaza_elo_serii(grp) for _, grp in df.groupby("Liga", observed=True)]
    df = pd.concat(elo_parts).sort_values(["Liga", "Date"]).reset_index(drop=True)

    df["home_goals_scored"] = rolling_ewm(df, "HomeTeam", "FTHG", decay=0.05)
    df["away_goals_scored"] = rolling_ewm(df, "AwayTeam", "FTAG", decay=0.05)
    df["home_goals_conceded"] = rolling_ewm(df, "HomeTeam", "FTAG", decay=0.05)
    df["away_goals_conceded"] = rolling_ewm(df, "AwayTeam", "FTHG", decay=0.05)

    xg_disponibil = False
    if {"HST", "AST"}.issubset(df.columns):
        xg_disponibil = True
        df["home_sot"] = rolling_ewm(df, "HomeTeam", "HST", decay=0.05)
        df["away_sot"] = rolling_ewm(df, "AwayTeam", "AST", decay=0.05)
        df["home_sot_conceded"] = rolling_ewm(df, "HomeTeam", "AST", decay=0.05)
        df["away_sot_conceded"] = rolling_ewm(df, "AwayTeam", "HST", decay=0.05)

        dfd = df.sort_values("Date").copy()
        dfd["league_goals_cum"] = dfd.groupby("Liga", observed=True)["FTHG"].cumsum().shift(1)
        dfd["league_sot_cum"] = dfd.groupby("Liga", observed=True)["HST"].cumsum().shift(1)
        dfd["conv_league_raw"] = dfd["league_goals_cum"] / (dfd["league_sot_cum"] + 1e-6)
        dfd["conv_league"] = (0.7 * dfd["conv_league_raw"] + 0.3 * 0.30).fillna(0.30)
        df = dfd.sort_index()

        df["home_xg_proxy"] = df["home_sot"] * df["conv_league"]
        df["away_xg_proxy"] = df["away_sot"] * df["conv_league"]
        df["home_xg_conceded_proxy"] = df["home_sot_conceded"] * df["conv_league"]
        df["away_xg_conceded_proxy"] = df["away_sot_conceded"] * df["conv_league"]
        df["xg_diff"] = df["home_xg_proxy"] - df["away_xg_proxy"]

        home_sot_raw = rolling_ewm(df, "HomeTeam", "HST", decay=0.03)
        home_goals = rolling_ewm(df, "HomeTeam", "FTHG", decay=0.03)
        away_sot_raw = rolling_ewm(df, "AwayTeam", "AST", decay=0.03)
        away_goals = rolling_ewm(df, "AwayTeam", "FTAG", decay=0.03)
        df["home_conversion"] = home_goals / (home_sot_raw + 0.1)
        df["away_conversion"] = away_goals / (away_sot_raw + 0.1)
        df["conversion_diff"] = df["home_conversion"] - df["away_conversion"]
        df["home_xg_proxy"] *= (0.7 + 0.6 * df["home_conversion"].clip(0.2, 0.6))
        df["away_xg_proxy"] *= (0.7 + 0.6 * df["away_conversion"].clip(0.2, 0.6))

        if {"HS", "AS"}.issubset(df.columns):
            df["home_shots"] = rolling_ewm(df, "HomeTeam", "HS", decay=0.05)
            df["away_shots"] = rolling_ewm(df, "AwayTeam", "AS", decay=0.05)
            df["sot_pct_home"] = df["home_sot"] / (df["home_shots"] + 0.1)
            df["sot_pct_away"] = df["away_sot"] / (df["away_shots"] + 0.1)
    else:
        df["home_xg_proxy"] = df["home_goals_scored"]
        df["away_xg_proxy"] = df["away_goals_scored"]
        df["home_xg_conceded_proxy"] = df["home_goals_conceded"]
        df["away_xg_conceded_proxy"] = df["away_goals_conceded"]
        df["xg_diff"] = df["home_xg_proxy"] - df["away_xg_proxy"]
        for col in ["home_sot", "away_sot", "home_sot_conceded", "away_sot_conceded", "home_conversion", "away_conversion", "conversion_diff", "sot_pct_home", "sot_pct_away"]:
            df[col] = np.nan

    df.attrs["xg_din_suturi"] = xg_disponibil

    for liga_name, idx in df.groupby("Liga", observed=True).groups.items():
        df_l = df.loc[idx]
        mu_scored = max(df_l["home_xg_proxy"].mean(), 0.3)
        mu_conceded = max(df_l["away_xg_conceded_proxy"].mean(), 0.3)
        mu_league = max((mu_scored + mu_conceded) / 2, 0.3)
        atac_h = df_l["home_xg_proxy"] / (mu_scored + 1e-6)
        apar_a = df_l["away_xg_conceded_proxy"] / (mu_conceded + 1e-6)
        atac_a = df_l["away_xg_proxy"] / (mu_conceded + 1e-6)
        apar_h = df_l["home_xg_conceded_proxy"] / (mu_scored + 1e-6)
        df.loc[idx, "lambda_home"] = (atac_h * apar_a * mu_league).clip(0.2, 5.0).values
        df.loc[idx, "lambda_away"] = (atac_a * apar_h * mu_league).clip(0.2, 5.0).values
        df.loc[idx, "mu_league"] = mu_league

    df["lambda_total"] = df["lambda_home"] + df["lambda_away"]

    if {"HC", "AC"}.issubset(df.columns):
        df["home_corners"] = rolling_ewm(df, "HomeTeam", "HC", decay=0.05)
        df["away_corners"] = rolling_ewm(df, "AwayTeam", "AC", decay=0.05)
        df["corners_total"] = df["home_corners"] + df["away_corners"]
    else:
        df["home_corners"] = np.nan
        df["away_corners"] = np.nan
        df["corners_total"] = np.nan

    if "corners_total" in df.columns and "home_sot" in df.columns and "away_sot" in df.columns:
        df["league_tempo"] = (
            df.groupby("Liga", observed=True)["corners_total"].transform("mean") +
            df.groupby("Liga", observed=True)["home_sot"].transform("mean") +
            df.groupby("Liga", observed=True)["away_sot"].transform("mean")
        )
    else:
        df["league_tempo"] = np.nan

    if {"HF", "AF"}.issubset(df.columns):
        df["home_fouls"] = rolling_ewm(df, "HomeTeam", "HF", decay=0.05)
        df["away_fouls"] = rolling_ewm(df, "AwayTeam", "AF", decay=0.05)
    else:
        df["home_fouls"] = np.nan
        df["away_fouls"] = np.nan

    if {"HY", "AY"}.issubset(df.columns):
        df["home_yellows"] = rolling_ewm(df, "HomeTeam", "HY", decay=0.05)
        df["away_yellows"] = rolling_ewm(df, "AwayTeam", "AY", decay=0.05)
    else:
        df["home_yellows"] = np.nan
        df["away_yellows"] = np.nan

    df["elo_diff_norm"] = df["elo_diff"] / 400.0
    df["attack_h_vs_def_a"] = df["home_xg_proxy"] * df["away_xg_conceded_proxy"]
    df["attack_a_vs_def_h"] = df["away_xg_proxy"] * df["home_xg_conceded_proxy"]
    df["btts_proxy"] = df["lambda_home"] * df["lambda_away"]
    df["strength_diff"] = (
        (df["home_xg_proxy"] - df["home_xg_conceded_proxy"]) -
        (df["away_xg_proxy"] - df["away_xg_conceded_proxy"])
    )

    home_rates = df.groupby("HomeTeam", observed=True)["total_goals"].transform("mean")
    away_rates = df.groupby("AwayTeam", observed=True)["total_goals"].transform("mean")
    df["home_advantage"] = home_rates - away_rates

    df["season_progress"] = df.groupby(["Liga", "Sezon"], observed=True)["Date"].transform(
        lambda x: (x.rank(method="first") - 1) / max(len(x) - 1, 1)
    )

    df["home_days_rest"] = _days_rest(df, "HomeTeam").reindex(df.index).clip(3, 21)
    df["away_days_rest"] = _days_rest(df, "AwayTeam").reindex(df.index).clip(3, 21)
    df["rest_diff"] = df["home_days_rest"] - df["away_days_rest"]
    return df


@st.cache_data(show_spinner=False, ttl=3600)
def adauga_target_features_cached(data_hash: str, condition: str, df: pd.DataFrame) -> pd.DataFrame:
    return _adauga_target_features_core(df, condition)


def _adauga_target_features_core(df: pd.DataFrame, condition: str) -> pd.DataFrame:
    threshold = CONDITIONS[condition]["threshold"]
    df = df.copy()
    df["target"] = (df["total_goals"] > threshold).astype(int)
    df["home_over_rate"] = rolling_ewm(df, "HomeTeam", "target", decay=0.04)
    df["away_over_rate"] = rolling_ewm(df, "AwayTeam", "target", decay=0.04)
    league_mean = df.groupby("Liga", observed=True)["target"].transform("mean")
    df["home_over_rate_adj"] = df["home_over_rate"] - league_mean
    df["away_over_rate_adj"] = df["away_over_rate"] - league_mean
    df["league_over_mean"] = league_mean
    df["home_form5"] = rolling_ewm(df, "HomeTeam", "target", decay=0.25)
    df["away_form5"] = rolling_ewm(df, "AwayTeam", "target", decay=0.25)
    df["home_form10"] = rolling_ewm(df, "HomeTeam", "target", decay=0.12)
    df["away_form10"] = rolling_ewm(df, "AwayTeam", "target", decay=0.12)
    df["poisson_over"] = poisson_over_series(df["lambda_home"].values, df["lambda_away"].values, threshold)
    home_rates = df.groupby("HomeTeam", observed=True)["target"].transform("mean")
    away_rates = df.groupby("AwayTeam", observed=True)["target"].transform("mean")
    df["home_advantage"] = home_rates - away_rates
    return df


def antreneaza_model(df_base: pd.DataFrame, condition: str, use_optuna: bool = False):
    h = df_hash(df_base)
    df_cond = adauga_target_features_cached(h, condition, df_base)
    features_active = get_features_for_condition(condition, list(df_cond.columns))
    if not features_active:
        st.error(f"Niciun feature valid pentru {condition}.")
        return None

    df_sorted = df_cond.sort_values("Date").reset_index(drop=True)
    X = df_sorted[features_active].copy()
    y = df_sorted["target"].astype(int)
    col_means = X.mean(numeric_only=True)
    X = X.fillna(col_means)
    tscv = TimeSeriesSplit(n_splits=5)

    best_params = None
    if use_optuna:
        try:
            import optuna
            optuna.logging.set_verbosity(optuna.logging.WARNING)

            def objective(trial):
                params = {
                    "n_estimators": trial.suggest_int("n_estimators", 300, 800),
                    "max_depth": trial.suggest_int("max_depth", 3, 6),
                    "learning_rate": trial.suggest_float("learning_rate", 0.02, 0.08),
                    "subsample": trial.suggest_float("subsample", 0.65, 1.0),
                    "colsample_bytree": trial.suggest_float("colsample_bytree", 0.55, 1.0),
                    "reg_alpha": trial.suggest_float("reg_alpha", 0.0, 2.5),
                    "reg_lambda": trial.suggest_float("reg_lambda", 1.0, 6.0),
                    "min_child_weight": trial.suggest_int("min_child_weight", 5, 20),
                    "gamma": trial.suggest_float("gamma", 0.0, 0.5),
                }
                scores = []
                for train_idx, test_idx in tscv.split(X):
                    Xt, Xv = X.iloc[train_idx], X.iloc[test_idx]
                    yt, yv = y.iloc[train_idx], y.iloc[test_idx]
                    if len(yt) < 200 or len(yv) < 50:
                        continue
                    model = XGBClassifier(**params, random_state=42, eval_metric="logloss")
                    model.fit(Xt, yt, eval_set=[(Xv, yv)], verbose=False)
                    proba = model.predict_proba(Xv)[:, 1]
                    scores.append(roc_auc_score(yv, proba) - 0.15 * log_loss(yv, proba))
                return float(np.mean(scores)) if scores else 0.0

            study = optuna.create_study(direction="maximize")
            study.optimize(objective, n_trials=30, show_progress_bar=False)
            best_params = study.best_params
            st.info(f"[{condition}] Optuna score: {study.best_value:.3f}")
        except Exception:
            st.warning("Optuna indisponibil; se folosesc parametrii default.")

    params = best_params or DEFAULT_PARAMS[condition]
    auc_scores, brier_scores, all_probs, all_y = [], [], [], []

    for train_idx, test_idx in tscv.split(X):
        X_train, X_test = X.iloc[train_idx], X.iloc[test_idx]
        y_train, y_test = y.iloc[train_idx], y.iloc[test_idx]
        if len(y_train) < 200 or len(y_test) < 50:
            continue
        xgb = XGBClassifier(**params, random_state=42, eval_metric="logloss")
        cal = CalibratedClassifierCV(xgb, method="isotonic", cv=3)
        cal.fit(X_train, y_train)
        prob = cal.predict_proba(X_test)[:, 1]
        auc_scores.append(roc_auc_score(y_test, prob))
        brier_scores.append(brier_score_loss(y_test, prob))
        all_probs.extend(prob.tolist())
        all_y.extend(y_test.tolist())

    if not auc_scores:
        st.error(f"Date insuficiente pentru antrenare [{condition}].")
        return None

    final_xgb = XGBClassifier(**params, random_state=42, eval_metric="logloss")
    final_xgb.fit(X, y)
    final_model = CalibratedClassifierCV(final_xgb, method="isotonic", cv=3)
    final_model.fit(X, y)

    perm = permutation_importance(final_xgb, X, y, n_repeats=5, random_state=42, n_jobs=1)

    model_data = {
        "model": final_model,
        "features": features_active,
        "col_means": col_means.to_dict(),
        "params": params,
        "condition": condition,
        "threshold": CONDITIONS[condition]["threshold"],
        "auc": float(np.mean(auc_scores)),
        "auc_std": float(np.std(auc_scores)),
        "brier": float(np.mean(brier_scores)),
        "baseline_brier": float(y.mean() * (1 - y.mean())),
        "prob_all": all_probs,
        "y_all": all_y,
        "importances": perm.importances_mean.tolist(),
        "xg_din_suturi": df_base.attrs.get("xg_din_suturi", False),
        "trained_at": datetime.now().strftime("%d.%m.%Y %H:%M"),
        "model_version": datetime.now().strftime("%Y%m%d_%H%M"),
        "n_matches": int(len(df_cond)),
        "target_mean": float(y.mean()),
    }

    with open(MODEL_PATHS[condition], "wb") as f:
        pickle.dump(model_data, f)
    return model_data


def get_df_base_filtered(df_procesate: pd.DataFrame, sez_sel: list) -> pd.DataFrame:
    if df_procesate is None or df_procesate.empty:
        return pd.DataFrame()
    dff = df_procesate[df_procesate["Sezon"].isin(sez_sel)].copy()
    dff.attrs = dict(df_procesate.attrs)
    return dff


def get_df_cond(df_base: pd.DataFrame, condition: str) -> pd.DataFrame:
    if df_base.empty:
        return pd.DataFrame()
    return adauga_target_features_cached(df_hash(df_base), condition, df_base)


def _latest_home_away_rows(home: str, away: str, liga: str, df_cond: pd.DataFrame):
    df_liga = df_cond[df_cond["Liga"] == liga]
    home_rows = df_liga[df_liga["HomeTeam"] == home].sort_values("Date")
    away_rows = df_liga[df_liga["AwayTeam"] == away].sort_values("Date")
    if home_rows.empty or away_rows.empty:
        return df_liga, None, None
    return df_liga, home_rows.iloc[-1], away_rows.iloc[-1]


def _build_match_context(home: str, away: str, liga: str, df_cond: pd.DataFrame, threshold: float):
    df_liga, hl, al = _latest_home_away_rows(home, away, liga, df_cond)
    if hl is None or al is None:
        return None
    mu_league = sfloat(df_liga["mu_league"].mean() if "mu_league" in df_liga.columns else np.nan, 1.2)
    mu_scored = sfloat(df_liga["home_xg_proxy"].mean(), 1.0)
    mu_conceded = sfloat(df_liga["away_xg_conceded_proxy"].mean(), 1.0)
    hxg_sc = sfloat(hl.get("home_xg_proxy", np.nan), 1.0)
    axg_sc = sfloat(al.get("away_xg_proxy", np.nan), 1.0)
    hxg_con = sfloat(hl.get("home_xg_conceded_proxy", np.nan), 1.0)
    axg_con = sfloat(al.get("away_xg_conceded_proxy", np.nan), 1.0)
    atac_h = hxg_sc / (mu_scored + 1e-6)
    apar_a = axg_con / (mu_conceded + 1e-6)
    atac_a = axg_sc / (mu_conceded + 1e-6)
    apar_h = hxg_con / (mu_scored + 1e-6)
    lh = float(np.clip(atac_h * apar_a * mu_league, 0.2, 5.0))
    la = float(np.clip(atac_a * apar_h * mu_league, 0.2, 5.0))
    rata_h = sfloat(hl.get("home_over_rate", np.nan))
    rata_a = sfloat(al.get("away_over_rate", np.nan))
    rata_medie = (rata_h + rata_a) / 2 if not np.isnan(rata_h) and not np.isnan(rata_a) else np.nan
    poisson_p = poisson_over(lh, la, threshold)
    return {
        "df_liga": df_liga,
        "hl": hl,
        "al": al,
        "lambda_home": lh,
        "lambda_away": la,
        "lambda_total": lh + la,
        "poisson_p": poisson_p,
        "rata_medie": rata_medie,
        "discrepanta": abs(poisson_p - rata_medie) if not np.isnan(poisson_p) and not np.isnan(rata_medie) else np.nan,
        "hxg_sc": hxg_sc,
        "axg_sc": axg_sc,
        "hxg_con": hxg_con,
        "axg_con": axg_con,
    }


def construieste_vector(home: str, away: str, liga: str, df_cond: pd.DataFrame, model_data: dict):
    features = model_data["features"]
    col_means = model_data["col_means"]
    condition = model_data["condition"]
    threshold = model_data["threshold"]
    ctx = _build_match_context(home, away, liga, df_cond, threshold)
    if ctx is None:
        return None
    hl, al, df_liga = ctx["hl"], ctx["al"], ctx["df_liga"]
    data = {}
    for f in features:
        if f.startswith("home_") or f == "elo_home":
            data[f] = hl.get(f, np.nan)
        elif f.startswith("away_") or f == "elo_away":
            data[f] = al.get(f, np.nan)
        else:
            data[f] = np.nan
    derived = {
        "lambda_home": ctx["lambda_home"],
        "lambda_away": ctx["lambda_away"],
        "lambda_total": ctx["lambda_total"],
        "poisson_over": ctx["poisson_p"],
        "elo_diff": sfloat(hl.get("elo_home", 1500), 1500) - sfloat(al.get("elo_away", 1500), 1500),
        "elo_diff_norm": (sfloat(hl.get("elo_home", 1500), 1500) - sfloat(al.get("elo_away", 1500), 1500)) / 400.0,
        "attack_h_vs_def_a": ctx["hxg_sc"] * ctx["axg_con"],
        "attack_a_vs_def_h": ctx["axg_sc"] * ctx["hxg_con"],
        "btts_proxy": ctx["lambda_home"] * ctx["lambda_away"],
        "strength_diff": (ctx["hxg_sc"] - ctx["hxg_con"]) - (ctx["axg_sc"] - ctx["axg_con"]),
        "league_over_mean": df_liga["target"].mean() if "target" in df_liga.columns else np.nan,
        "season_progress": sfloat(hl.get("season_progress", 0.5), 0.5),
        "rest_diff": 0.0,
        "league_tempo": sfloat(hl.get("league_tempo", np.nan)),
    }
    for key, val in derived.items():
        if key in features:
            data[key] = val
    X_new = pd.DataFrame([data])[features]
    fill_vals = {col: col_means.get(col, 0.0) for col in X_new.columns}
    return X_new.fillna(fill_vals)


def get_prob(home: str, away: str, liga: str, df_cond: pd.DataFrame, model_data: dict):
    X = construieste_vector(home, away, liga, df_cond, model_data)
    if X is None:
        return None
    return float(model_data["model"].predict_proba(X)[0, 1])


def get_diagnostics(home: str, away: str, liga: str, df_cond: pd.DataFrame, model_data: dict):
    ctx = _build_match_context(home, away, liga, df_cond, model_data["threshold"])
    if ctx is None:
        return {}
    hl, al = ctx["hl"], ctx["al"]

    def fmt(v, pct=False):
        if v is None or (isinstance(v, float) and np.isnan(v)):
            return "—"
        return f"{v:.1%}" if pct else round(v, 3)

    return {
        "lambda_home": ctx["lambda_home"],
        "lambda_away": ctx["lambda_away"],
        "lambda_total": ctx["lambda_total"],
        "poisson_p": ctx["poisson_p"],
        "home_sot": fmt(sfloat(hl.get("home_sot", np.nan))),
        "away_sot": fmt(sfloat(al.get("away_sot", np.nan))),
        "home_conversion": fmt(sfloat(hl.get("home_conversion", np.nan))),
        "away_conversion": fmt(sfloat(al.get("away_conversion", np.nan))),
        "home_over_rate": fmt(sfloat(hl.get("home_over_rate", np.nan)), pct=True),
        "away_over_rate": fmt(sfloat(al.get("away_over_rate", np.nan)), pct=True),
        "home_form5": fmt(sfloat(hl.get("home_form5", np.nan)), pct=True),
        "away_form5": fmt(sfloat(al.get("away_form5", np.nan)), pct=True),
        "elo_home": round(sfloat(hl.get("elo_home", 1500), 1500), 0),
        "elo_away": round(sfloat(al.get("elo_away", 1500), 1500), 0),
        "home_xg_proxy": fmt(ctx["hxg_sc"]),
        "away_xg_proxy": fmt(ctx["axg_sc"]),
        "discrepanta": ctx["discrepanta"],
        "rata_medie": ctx["rata_medie"],
    }


def calculeaza_edge(
    prob: float,
    cota_o: float,
    cota_u: float,
    bankroll: float = 100.0,
    fractie_kelly: float = 0.5
) -> dict:
    prob_u = 1.0 - prob

    raw_o = 1.0 / cota_o
    raw_u = 1.0 / cota_u
    margin = raw_o + raw_u
    bk_prob_o = raw_o / margin
    bk_prob_u = raw_u / margin

    edge_o = prob - bk_prob_o
    edge_u = prob_u - bk_prob_u

    # EV pe unitate mizată
    ev_o = prob * (cota_o - 1) - (1 - prob)
    ev_u = prob_u * (cota_u - 1) - (1 - prob_u)

    kelly_o = max(0.0, min((prob * cota_o - 1) / (cota_o - 1), 0.30))
    kelly_u = max(0.0, min((prob_u * cota_u - 1) / (cota_u - 1), 0.30))

    stake_o = bankroll * fractie_kelly * kelly_o
    stake_u = bankroll * fractie_kelly * kelly_u

    # alege selecția cu EV mai bun
    if ev_o >= ev_u:
        best_side = "Over"
        best_stake = stake_o
    else:
        best_side = "Under"
        best_stake = stake_u

    return {
        "prob_o": prob,
        "prob_u": prob_u,
        "bk_prob_o": bk_prob_o,
        "bk_prob_u": bk_prob_u,
        "margin": margin,
        "edge_o": edge_o,
        "edge_u": edge_u,
        "ev_o": ev_o,
        "ev_u": ev_u,
        "kelly_o": kelly_o,
        "kelly_u": kelly_u,
        "stake_o": stake_o,
        "stake_u": stake_u,
        "best_side": best_side,
        "best_stake": best_stake,
        "bankroll": bankroll,
        "fractie_kelly": fractie_kelly,
        "cota_fair_o": round(1.0 / prob, 2) if prob > 0.01 else 99.0,
        "cota_fair_u": round(1.0 / prob_u, 2) if prob_u > 0.01 else 99.0,
        "value_o": edge_o > 0.03 and ev_o > 0,
        "value_u": edge_u > 0.03 and ev_u > 0,
    }

def verdict_html(ed: dict, label: str = "Over") -> str:
    if ed["value_o"] and ed["value_u"]:
        best = label if ed["ev_o"] >= ed["ev_u"] else "Under"
        color, border = "#198754", "#198754"
        txt = (
            f"✅ VALUE BET — {best} | "
            f"EV Over: {ed['ev_o']:+.2f}u | EV Under: {ed['ev_u']:+.2f}u | "
            f"Miză: {ed['best_stake']:.2f}u ({ed['best_side']})"
        )
    elif ed["value_o"]:
        color, border = "#198754", "#198754"
        txt = (
            f"✅ VALUE BET — {label} | "
            f"EV: {ed['ev_o']:+.2f}u | "
            f"Kelly: {ed['kelly_o']*100:.1f}% | "
            f"Miză: {ed['stake_o']:.2f}u | "
            f"Edge: {ed['edge_o']*100:+.1f}%"
        )
    elif ed["value_u"]:
        color, border = "#198754", "#198754"
        txt = (
            f"✅ VALUE BET — Under | "
            f"EV: {ed['ev_u']:+.2f}u | "
            f"Kelly: {ed['kelly_u']*100:.1f}% | "
            f"Miză: {ed['stake_u']:.2f}u | "
            f"Edge: {ed['edge_u']*100:+.1f}%"
        )
    elif ed["ev_o"] > 0 or ed["ev_u"] > 0:
        color, border = "#b45309", "#ffc107"
        txt = "⚠️ EV pozitiv — edge mic (sub 3%) — posibil noise"
    else:
        color, border = "#dc3545", "#dc3545"
        txt = "❌ Fără value — BK acoperit"

    return f"""
    <div style="background:#fff;border:2px solid {border};border-radius:10px;
        padding:1rem 1.4rem;margin:0.5rem 0">
        <span style="color:{color};font-size:20px;font-weight:700">{txt}</span>
    </div>
    """

def get_history_columns():
    return [
        "ts", "conditie", "home", "away", "liga",
        "prob_o", "prob_u", "cota_o", "cota_u",
        "edge_o", "edge_u", "ev_o", "ev_u",
        "kelly_o", "kelly_u", "stake_o", "stake_u",
        "best_side", "best_stake", "bankroll", "fractie_kelly",
        "value", "bettable", "decision"
    ]


def _style_history_sheet(ws):
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = True

    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    widths = {
        "A": 18, "B": 18, "C": 18, "D": 18, "E": 20,
        "F": 10, "G": 10, "H": 9, "I": 9,
        "J": 10, "K": 10, "L": 9, "M": 9,
        "N": 10, "O": 10, "P": 10, "Q": 10,
        "R": 12, "S": 12, "T": 10, "U": 12,
        "V": 10, "W": 10, "X": 10
    }

    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    last_row = ws.max_row
    last_col = ws.max_column
    if last_row >= 1 and last_col >= 1:
        ws.auto_filter.ref = f"A1:{get_column_letter(last_col)}{last_row}"

def _refresh_history_table(ws):
    if ws.max_row < 2 or ws.max_column < 1:
        return

    ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"

    existing_tables = list(ws.tables.keys())
    for name in existing_tables:
        del ws.tables[name]

    tab = Table(displayName=HISTORY_TABLE, ref=ref)
    style = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    tab.tableStyleInfo = style
    ws.add_table(tab)

def salveaza_predictie(home, away, liga, condition, prob, ed, cota_o, cota_u):
    columns = get_history_columns()

    record = {
        "ts": datetime.now().strftime("%d.%m.%Y %H:%M"),
        "conditie": condition,
        "home": home,
        "away": away,
        "liga": liga,
        "prob_o": round(prob, 4),
        "prob_u": round(1 - prob, 4),
        "cota_o": cota_o,
        "cota_u": cota_u,
        "edge_o": round(ed["edge_o"], 4),
        "edge_u": round(ed["edge_u"], 4),
        "ev_o": round(ed["ev_o"], 3),
        "ev_u": round(ed["ev_u"], 3),
        "kelly_o": round(ed["kelly_o"], 4),
        "kelly_u": round(ed["kelly_u"], 4),
        "stake_o": round(ed.get("stake_o", 0.0), 2),
        "stake_u": round(ed.get("stake_u", 0.0), 2),
        "best_side": ed.get("best_side", ""),
        "best_stake": round(ed.get("best_stake", 0.0), 2),
        "bankroll": round(ed.get("bankroll", 0.0), 2),
        "fractie_kelly": round(ed.get("fractie_kelly", 0.0), 4),
        "value": bool(ed["value_o"] or ed["value_u"]),
        "bettable": bool(ed.get("best_stake", 0.0) > 0),
        "decision": "BET" if ed.get("best_stake", 0.0) > 0 else "PASS",
    }

    if not os.path.exists(HISTORY_PATH):
        wb = Workbook()
        ws = wb.active
        ws.title = HISTORY_SHEET
        ws.append(columns)
        ws.append([record.get(col, "") for col in columns])
        _style_history_sheet(ws)
        _refresh_history_table(ws)
        wb.save(HISTORY_PATH)
        wb.close()
        return

    wb = load_workbook(HISTORY_PATH)
    if HISTORY_SHEET in wb.sheetnames:
        ws = wb[HISTORY_SHEET]
    else:
        ws = wb.active
        ws.title = HISTORY_SHEET
        if ws.max_row == 1 and ws["A1"].value is None:
            ws.delete_rows(1, 1)
        ws.append(columns)

    # dacă headerul lipsește, îl reconstruiește
    first_row = [ws.cell(row=1, column=i).value for i in range(1, len(columns) + 1)]
    if first_row != columns:
        ws.delete_rows(1, ws.max_row)
        ws.append(columns)

    ws.append([record.get(col, "") for col in columns])

    _style_history_sheet(ws)
    _refresh_history_table(ws)

    wb.save(HISTORY_PATH)
    wb.close()



with st.sidebar:
    st.title("Predictii O/U")
    st.markdown("""
    <div style="
        background-color:#dbeafe;
        color:#0f172a;
        padding:12px 14px;
        border-radius:10px;
        border:1px solid #93c5fd;
    ">
        <strong style="color:#1e3a8a;">3 modele independente</strong><br>
        <span style="color:#0f172a;">
            Over/Under 1.5 · 2.5 · 3.5<br>
            Antrenate pe 10+ sezoane, cu 100+ features
        </span>
    </div>
    """, unsafe_allow_html=True)

    condition_sel = st.radio(
        "Selectează condiția",
        list(CONDITIONS.keys()),
        format_func=lambda c: f"{CONDITIONS[c]['emoji']} {c}",
        key="condition_radio",
    )

    st.markdown("---")
    all_models = load_all_models()
    st.subheader("Status modele")
    for cond, md in all_models.items():
        if md:
            auc = md.get("auc", 0)
            icon = "🟢" if auc >= 0.62 else ("🟠" if auc >= 0.57 else "🔴")
            st.markdown(f"{CONDITIONS[cond]['emoji']} {cond} {icon} AUC {auc:.3f} ± {md.get('auc_std', 0):.3f}")
            st.caption(f"{md.get('trained_at', '?')} • {md.get('n_matches', 0)} meciuri")
        else:
            st.markdown(f"{CONDITIONS[cond]['emoji']} {cond} — neantrenat")

    df_procesate = None
    if os.path.exists(DATA_PATH):
        try:
            with open(DATA_PATH, "rb") as f:
                df_procesate = pickle.load(f)
            if not isinstance(df_procesate, pd.DataFrame) or df_procesate.empty:
                df_procesate = None
        except Exception:
            df_procesate = None

    st.markdown("---")
    st.subheader("Date")
    if df_procesate is not None:
        st.success(f"{len(df_procesate):,} meciuri procesate")
    else:
        st.warning("Date lipsă")

    sez_sel = st.multiselect("Sezoane incluse", SEZOANE, default=SEZOANE)

    st.markdown("---")
    st.subheader("Filtre value")
    prag_prob = st.slider("Prob. minimă model (%)", 20, 85, 50)
    prag_edge = st.slider("Edge minim (%)", 0, 15, 3)
    bankroll = st.number_input("Bankroll (u)", min_value=1.0, max_value=1_000_000.0, value=100.0, step=1.0)
    fractie_kelly = st.slider("Fracție Kelly", min_value=0.0, max_value=1.0, value=0.50, step=0.05)

    st.markdown("---")
    st.subheader("Antrenare")
    use_optuna = st.checkbox("Optimizare Optuna (mai lent)", value=False)
    btn_descarca = st.button("Descarcă și procesează date")
    btn_train_all = st.button("Antrenează toate modelele")
    btn_train_cond = st.button(f"Antrenează doar {condition_sel}")

if btn_descarca:
    with st.spinner("Descărcare date football-data.co.uk ..."):
        prog = st.progress(0.0)
        stxt = st.empty()
        df_raw = descarca_fdco(prog, stxt)
    if df_raw.empty:
        st.error("Nu s-au putut descărca datele.")
    else:
        with st.spinner("Procesare features de bază ..."):
            raw_h = df_hash(df_raw)
            df_procesate = proceseaza_date_cached(raw_h, df_raw)
            df_procesate.attrs["xg_din_suturi"] = df_raw.attrs.get("xg_din_suturi", False)
            with open(DATA_PATH, "wb") as f:
                pickle.dump(df_procesate, f)
        st.success(f"{len(df_procesate):,} meciuri procesate și salvate.")
        st.rerun()

if df_procesate is None:
    df_base = pd.DataFrame()
else:
    df_base = get_df_base_filtered(df_procesate, sez_sel)

if (btn_train_all or btn_train_cond) and not df_base.empty:
    conds = list(CONDITIONS.keys()) if btn_train_all else [condition_sel]
    for cond in conds:
        with st.spinner(f"Antrenare {cond} ..."):
            md = antreneaza_model(df_base, cond, use_optuna=use_optuna)
            if md:
                st.success(f"{cond} — AUC {md['auc']:.3f} ± {md['auc_std']:.3f} | Brier {md['brier']:.3f}")
    st.rerun()

model_activ = all_models.get(condition_sel)
df_cond = get_df_cond(df_base, condition_sel) if not df_base.empty else pd.DataFrame()
cond_css = {"Over/Under 1.5": "cond-15", "Over/Under 2.5": "cond-25", "Over/Under 3.5": "cond-35"}
st.markdown(f"<span class='condition-badge {cond_css[condition_sel]}'>{CONDITIONS[condition_sel]['emoji']} {condition_sel} activ</span>", unsafe_allow_html=True)

tab1, tab2, tab3, tab4, tab5 = st.tabs(["Predicție", "Model", "Backtesting", "Comparare", "Istoric"])

if "last_prediction" not in st.session_state:
    st.session_state.last_prediction = None

with tab1:
    if model_activ is None:
        st.warning(f"⚠️ Modelul pentru {condition_sel} nu a fost antrenat.")
    elif df_cond.empty:
        st.warning("⚠️ Date lipsă — descarcă și procesează datele mai întâi.")
    else:
        ligi_disponibile = sorted(df_cond["Liga"].astype(str).unique().tolist())
        c1, c2 = st.columns([1, 1.2])

        with c1:
            liga_sel = st.selectbox("🏆 Ligă", ligi_disponibile)
            df_liga = df_cond[df_cond["Liga"].astype(str) == liga_sel]
            echipe_h = sorted(df_liga["HomeTeam"].astype(str).unique().tolist())
            echipe_a = sorted(df_liga["AwayTeam"].astype(str).unique().tolist())
            home_sel = st.selectbox("🏠 Echipă gazdă", echipe_h)
            away_sel = st.selectbox("✈️ Echipă oaspete", echipe_a)

            st.markdown("**Cote bookmaker**")
            cota_o = st.number_input("Cotă Over", min_value=1.01, value=1.90, step=0.01)
            cota_u = st.number_input("Cotă Under", min_value=1.01, value=1.90, step=0.01)

            btn_pred = st.button("🔮 Calculează predicția", type="primary", use_container_width=True)

        # 1) Când apeși Calculează, computezi și scrii în session_state
        if btn_pred:
            with c2:
                if home_sel == away_sel:
                    st.error("Echipa gazdă și oaspete trebuie să fie diferite.")
                else:
                    prob = get_prob(home_sel, away_sel, liga_sel, df_liga, model_activ)
                    if prob is None:
                        st.error("Date insuficiente pentru aceste echipe.")
                    else:
                        diag = get_diagnostics(home_sel, away_sel, liga_sel, df_liga, model_activ)
                        ed = calculeaza_edge(
                            prob,
                            cota_o,
                            cota_u,
                            bankroll=bankroll,
                            fractie_kelly=fractie_kelly,
                        )
                        ci_lo, ci_hi = bootstrap_ci(prob)

                        st.session_state.last_prediction = {
                            "home": home_sel,
                            "away": away_sel,
                            "liga": liga_sel,
                            "condition": condition_sel,
                            "prob": prob,
                            "diag": diag,
                            "ed": ed,
                            "cota_o": cota_o,
                            "cota_u": cota_u,
                            "ci_lo": ci_lo,
                            "ci_hi": ci_hi,
                            "bankroll": bankroll,
                            "fractie_kelly": fractie_kelly,
                        }

        # 2) Afișare din session_state
        pred = st.session_state.last_prediction

        with c2:
            if pred is not None and pred["condition"] == condition_sel:
                prob = pred["prob"]
                ed = pred["ed"]
                diag = pred["diag"]
                ci_lo = pred["ci_lo"]
                ci_hi = pred["ci_hi"]
                bankroll = pred["bankroll"]
                fractie_kelly = pred["fractie_kelly"]

                st.markdown(f"### {pred['home']} vs {pred['away']}")
                st.markdown(f"**Ligă:** {pred['liga']} | **Condiție:** {pred['condition']}")

                m1, m2, m3, m4, m5 = st.columns(5)
                m1.metric("Prob. Over", f"{prob:.1%}")
                m2.metric("Prob. Under", f"{1-prob:.1%}")
                m3.metric("Edge Over", f"{ed['edge_o']*100:+.1f}%")
                m4.metric("EV Over", f"{ed['ev_o']:+.2f}u")
                m5.metric("Miză rec.", f"{ed['best_stake']:.2f}u", ed["best_side"])

                st.caption(
                    f"CI 90%: {ci_lo:.1%} – {ci_hi:.1%} | "
                    f"Fair Over {ed['cota_fair_o']:.2f} | Fair Under {ed['cota_fair_u']:.2f}"
                )
                st.markdown(
                    verdict_html(ed, f"Over {CONDITIONS[condition_sel]['label']}"),
                    unsafe_allow_html=True,
                )

                d1, d2, d3 = st.columns(3)
                d1.metric("Bankroll", f"{bankroll:.2f}u")
                d2.metric("Fracție Kelly", f"{fractie_kelly:.0%}")
                d3.metric("Miza selecției", f"{ed['best_stake']:.2f}u", ed["best_side"])

                with st.expander("🔍 Diagnostice detaliate"):
                    dc1, dc2 = st.columns(2)
                    with dc1:
                        st.markdown(f"**Lambda gazdă:** {diag['lambda_home']:.2f}")
                        st.markdown(f"**Lambda oaspeți:** {diag['lambda_away']:.2f}")
                        st.markdown(f"**Lambda total:** {diag['lambda_total']:.2f}")
                        st.markdown(f"**Poisson Over:** {diag['poisson_p']:.1%}")
                        st.markdown(
                            f"**Rată medie echipe:** {diag['rata_medie']:.1%}"
                            if diag.get("rata_medie") is not None and not np.isnan(diag["rata_medie"])
                            else "**Rată medie echipe:** —"
                        )
                    with dc2:
                        st.markdown(f"**SoT gazdă:** {diag['home_sot']}")
                        st.markdown(f"**SoT oaspeți:** {diag['away_sot']}")
                        st.markdown(f"**Conv. gazdă:** {diag['home_conversion']}")
                        st.markdown(f"**Conv. oaspeți:** {diag['away_conversion']}")
                        st.markdown(f"**Elo gazdă:** {diag['elo_home']}")
                        st.markdown(f"**Elo oaspeți:** {diag['elo_away']}")
                        st.markdown(f"**Form5 gazdă:** {diag['home_form5']}")
                        st.markdown(f"**Form5 oaspeți:** {diag['away_form5']}")

                if st.button("💾 Salvează predicția în istoric"):
                    salveaza_predictie(home_sel, away_sel, liga_sel, condition_sel, prob, ed, cota_o, cota_u)
                    st.success("Predicție salvată în Predictii_OU.xlsx")
            else:
                st.info("Calculează o predicție pentru a vedea detaliile și a o salva.")

with tab2:
    if model_activ is None:
        st.warning(f"Modelul pentru {condition_sel} nu a fost antrenat.")
    else:
        md = model_activ
        m1, m2, m3, m4 = st.columns(4)
        m1.metric("AUC CV", f"{md['auc']:.3f}", f"± {md['auc_std']:.3f}")
        m2.metric("Brier", f"{md['brier']:.3f}", f"baseline {md['baseline_brier']:.3f}")
        m3.metric("Meciuri", f"{md['n_matches']:,}")
        m4.metric("Rate Over", f"{md['target_mean']:.1%}")
        st.caption(f"Antrenat {md.get('trained_at', '?')} | Versiune {md.get('model_version', '?')}")
        if md.get("xg_din_suturi"):
            st.markdown("<span class='badge badge-green'>xG din șuturi pe cadru</span>", unsafe_allow_html=True)
        else:
            st.markdown("<span class='badge badge-orange'>xG fallback din goluri</span>", unsafe_allow_html=True)

        prob_all = md.get("prob_all", [])
        y_all = md.get("y_all", [])
        if prob_all and y_all:
            frac_pos, mean_pred = calibration_curve(y_all, prob_all, n_bins=10)
            fig = go.Figure()
            fig.add_trace(go.Scatter(x=mean_pred, y=frac_pos, mode="lines+markers", name="Model"))
            fig.add_trace(go.Scatter(x=[0, 1], y=[0, 1], mode="lines", name="Perfect", line=dict(dash="dash", color="gray")))
            fig.update_layout(height=360, margin=dict(l=40, r=20, t=20, b=40), xaxis_title="Probabilitate medie prezisă", yaxis_title="Fracție pozitivă reală")
            st.plotly_chart(fig, width='stretch')

with tab3:
    if model_activ is None:
        st.warning(f"Modelul pentru {condition_sel} nu a fost antrenat.")
    else:
        prob_all = model_activ.get("prob_all", [])
        y_all = model_activ.get("y_all", [])
        if not prob_all:
            st.info("Rulează antrenamentul pentru a genera date de backtesting.")
        else:
            prob_arr = np.array(prob_all)
            y_arr = np.array(y_all)
            prag_bt = st.slider("Prag probabilitate backtesting (%)", 40, 80, 55) / 100
            mask = prob_arr >= prag_bt
            n_pariuri = int(mask.sum())
            if n_pariuri == 0:
                st.warning("Niciun pariu la pragul selectat.")
            else:
                castiguri = int(y_arr[mask].sum())
                rata = castiguri / n_pariuri
                b1, b2, b3 = st.columns(3)
                b1.metric("Pariuri", f"{n_pariuri}")
                b2.metric("Câștiguri", f"{castiguri}")
                b3.metric("Rată câștig", f"{rata:.1%}")

with tab4:
    if df_base.empty:
        st.warning("Date lipsă.")
    else:
        ligi_comp = sorted(df_base["Liga"].astype(str).unique().tolist())
        c1, c2, c3 = st.columns(3)
        liga_comp = c1.selectbox("Ligă", ligi_comp, key="liga_comp")
        df_liga = df_base[df_base["Liga"].astype(str) == liga_comp]
        teams = sorted(df_liga["HomeTeam"].astype(str).unique().tolist())
        home_comp = c2.selectbox("Gazdă", teams, key="home_comp")
        away_comp = c3.selectbox("Oaspeți", teams, key="away_comp")
        if st.button("Compară toate condițiile", type="primary"):
            if home_comp == away_comp:
                st.error("Echipele trebuie să fie diferite.")
            else:
                rezultate = []
                for cond in CONDITIONS:
                    md = all_models.get(cond)
                    if md is None:
                        rezultate.append({"Condiție": cond, "Prob. Over": "—", "Prob. Under": "—", "Poisson": "—", "Status": "Neantrenat"})
                        continue
                    dfc = get_df_cond(df_base, cond)
                    df_lc = dfc[dfc["Liga"].astype(str) == liga_comp]
                    p = get_prob(home_comp, away_comp, liga_comp, df_lc, md)
                    if p is None:
                        rezultate.append({"Condiție": cond, "Prob. Over": "—", "Prob. Under": "—", "Poisson": "—", "Status": "Date insuficiente"})
                    else:
                        diag = get_diagnostics(home_comp, away_comp, liga_comp, df_lc, md)
                        rezultate.append({
                            "Condiție": cond,
                            "Prob. Over": f"{p:.1%}",
                            "Prob. Under": f"{1-p:.1%}",
                            "Poisson": f"{diag.get('poisson_p', np.nan):.1%}" if diag.get("poisson_p") is not None and not np.isnan(diag.get("poisson_p", np.nan)) else "—",
                            "Lambda total": f"{diag.get('lambda_total', np.nan):.2f}" if diag else "—",
                            "Status": "OK",
                        })
                st.dataframe(pd.DataFrame(rezultate), width='stretch')

with tab5:
    st.subheader("Istoric predicții salvate")

    columns = get_history_columns()

    if not os.path.exists(HISTORY_PATH):
        st.info("Nicio predicție salvată.")
    else:
        try:
            df_hist = pd.read_excel(HISTORY_PATH, sheet_name=HISTORY_SHEET)

            if df_hist.empty:
                st.info("Istoricul este gol.")
            else:
                for col in columns:
                    if col not in df_hist.columns:
                        df_hist[col] = ""

                df_hist = df_hist[columns].copy()

                c1, c2, c3 = st.columns(3)
                cond_filter = c1.multiselect(
                    "Filtrează condiție",
                    list(CONDITIONS.keys()),
                    default=list(CONDITIONS.keys())
                )
                value_only = c2.checkbox("Doar value bets", value=False)
                bet_only = c3.checkbox("Doar BET", value=False)

                df_show = df_hist[df_hist["conditie"].isin(cond_filter)].copy()

                if value_only and "value" in df_show.columns:
                    df_show = df_show[df_show["value"] == True]

                if bet_only and "decision" in df_show.columns:
                    df_show = df_show[df_show["decision"] == "BET"]

                st.dataframe(df_show, width="stretch")

                d1, d2, d3 = st.columns(3)

                if d1.button("Șterge istoricul"):
                    os.remove(HISTORY_PATH)
                    st.success("Istoric șters.")
                    st.rerun()

                csv_data = df_show.to_csv(index=False).encode("utf-8-sig")
                d2.download_button(
                    "Export CSV",
                    csv_data,
                    file_name="Predictii_OU_export.csv",
                    mime="text/csv"
                )

                with open(HISTORY_PATH, "rb") as f:
                    d3.download_button(
                        "Descarcă o copie XLSX",
                        data=f,
                        file_name="Predictii_OU.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )

        except Exception as e:
            st.error(f"Eroare la citirea istoricului: {e}")